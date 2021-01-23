#-*-coding:utf-8-*-
import discord
import asyncio
import datetime 
import openpyxl

from beautifultable import BeautifulTable
from discord.ext import commands

client = discord.Client()

@client.event
async def on_ready():
	print("Bot Start")


@client.event
async def on_message(message): 
    
    if message.content.startswith('$'):
        if message.content.startswith('$r'):
            
            dinoIn=True
            cntIn=True
            insertIn=True
             
            #등록
            try :
                tmp = message.content.split(" ")
                workbook = openpyxl.load_workbook('C:/file/RequestList.xlsx')
                bdlist = workbook[tmp[1].title()]
            except :
                embed = discord.Embed(title="Warning",description="해당하는 공룡이 존재 하지않습니다", color=0xFF0000)
                await message.channel.send(embed=embed)
                dinoIn=False

            if (dinoIn) :
                try :
                    tmp = message.content.split(" ")
                    cnt = int(tmp[2])
                except :
                    embed = discord.Embed(title="Warning",description="숫자만넣어주세요",  color=0xFF0000)
                    await message.channel.send(embed=embed)
                    cntIn=False
            if(dinoIn and cntIn) :
                inCnt = 1
                for row in bdlist.rows:
                     noTmp = "B" + str(inCnt)
                     if(str(bdlist[noTmp].value) == str(message.author) and insertIn) :
                         embed = discord.Embed(title="Warning",description="이미 신청되어있습니다",  color=0xFF0000)
                         await message.channel.send(embed=embed)
                         insertIn = False
                     inCnt = inCnt+ 1

            if (cnt>=1 and cnt <= 30 and dinoIn and cntIn and insertIn) :
                today = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                uesrid = str(message.author)
                

                maxRow = 0
                try :
                    for row in bdlist.rows:
                        maxRow = maxRow+1
                except :
                	maxRow = 0
                print(maxRow)
                inCnt = maxRow + 1
                noTmp = "A" + str(inCnt)
                userIdTmp = "B" + str(inCnt)
                countTmp = "C" + str(inCnt)
                dateTmp = "D" + str(inCnt)

                bdlist[noTmp] = inCnt
                bdlist[userIdTmp] = uesrid
                bdlist[countTmp] = cnt 
                bdlist[dateTmp] = today
                workbook.save('C:/file/RequestList.xlsx')
                workbook.close()

                embed = discord.Embed(title=tmp[1].title()+"RequestSuccess",  color=0x00ff00)
                embed.add_field(name="No", value=inCnt, inline=True)
                embed.add_field(name="UserId", value=uesrid, inline=True)
                if (cnt>=1 and cnt <= 9) :
                    embed.add_field(name="Count　　　Date", value= str(cnt)+"　　　　　"+today, inline=True)
                else :
                    embed.add_field(name="Count　　　Date", value= str(cnt)+"　　　　  "+today, inline=True)

                await message.channel.send(embed=embed)
            elif (cnt<=0 and cnt >= 31 and dinoIn and cntIn and insertIn) :
                embed = discord.Embed(title="Warning",description="1~30까지만 입력가능",  color=0xFF0000)
                await message.channel.send(embed=embed)
            print(message.content)
        #브리더리스트
        elif (message.content == "$list"):
            workbook = openpyxl.load_workbook('C:/file/BreederList.xlsx')
            bdlist = workbook['BreederList']
            embed = discord.Embed(title="BreederList",  color=0x00ff00)
            all_values = []
            for row in bdlist.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                    if len(row_value) == 2 :
                        embed.add_field(name="Dino"+"　　　　"+"InCharge", value=str(row_value[0])+"　　　　"+str(row_value[1]), inline=False)
                all_values.append(row_value)
            print(all_values)
            
            await message.channel.send(embed=embed)
        #신청리스트
        elif message.content.startswith('$list'):
            try :
                tmp = message.content.split(" ")
                workbook = openpyxl.load_workbook('C:/file/RequestList.xlsx')
                bdlist = workbook[tmp[1].title()]
            except :
                embed = discord.Embed(title="Warning",description="해당하는공룡이 존재 하지않습니다",  color=0xFF0000)
                await message.channel.send(embed=embed)

            embed = discord.Embed(title=tmp[1].title()+"RequestList",  color=0x00ff00)
            all_values = []
            listCnt = 0
            for row in bdlist.rows:
                row_value = []
                for cell in row:
                    row_value.append(cell.value)
                    if len(row_value) == 4 :
                        listCnt = listCnt + 1 
                        if (listCnt>=1 and listCnt <= 8) :
                            embed.add_field(name="No    UserId", value=str(row_value[0]) +" 　 "+str(row_value[1]), inline=True)
                            embed.add_field(name="Count", value=str(row_value[2]), inline=True)
                            embed.add_field(name="Date", value=str(row_value[3]), inline=True)
                            if(listCnt == 8) :
                               await message.channel.send(embed=embed)
                        else :
                            listCnt = 1
                            embed = discord.Embed(color=0x00ff00)
                            embed.add_field(name="No    UserId", value=str(row_value[0]) +" 　 "+str(row_value[1]), inline=True)
                            embed.add_field(name="Count", value=str(row_value[2]), inline=True)
                            embed.add_field(name="Date", value=str(row_value[3]), inline=True)

                all_values.append(row_value)
       	    if (listCnt>=1 and listCnt <= 7) :
                await message.channel.send(embed=embed)
            print(all_values)
            
        #삭제
        elif message.content.startswith('$done'):
            dinoIn=True
            cntIn=True
            rolesIn=False

            try :
                tmp = message.content.split(" ")
                workbook = openpyxl.load_workbook('C:/file/RequestList.xlsx')
                bdlist = workbook[tmp[1].title()]
            except :
                embed = discord.Embed(title="Warning",description="해당하는공룡이 존재 하지않습니다",  color=0xFF0000)
                await message.channel.send(embed=embed)
                dinoIn = False
            

            if (dinoIn) :
                try :
                    tmp = message.content.split(" ")
                    no = int(tmp[2])
                except :
                    embed = discord.Embed(title="Warning",description="삭제하실 No를 작성해 주세요",  color=0xFF0000)
                    await message.channel.send(embed=embed)
                    cntIn=False

            
            for i in range(len(message.author.roles)):
                print(message.author.roles[i].name)
                if (str(message.author.roles[i].name)=="🦕Breeder") :
                    rolesIn=True
            if (not rolesIn) :
                embed = discord.Embed(title="Warning",description="🦕Breeder권한이 없습니다",  color=0xFF0000)
                await message.channel.send(embed=embed)


            if(dinoIn and cntIn and rolesIn) :

                bdlist.delete_rows(no) #no 행 삭제
                inCnt  = 1
                for row in bdlist.rows:
                    noTmp = "A" + str(inCnt)
                    bdlist[noTmp] = inCnt
                    inCnt = inCnt+ 1

                workbook.save('C:/file/RequestList.xlsx')
                workbook.close()
         
                embed = discord.Embed(title="No "+str(no)+" " +tmp[1].title()+"DoneSuccess",  color=0x00ff00)
                await message.channel.send(embed=embed)

        #도움말
        elif message.content.startswith('$help'):
            embed = discord.Embed(title="Command List",  color=0x00ff00)
            embed.add_field(name="$r dinoName count", value="공룡 신청 기능", inline=False)
            embed.add_field(name="$list", value="브리더 리스트를 확인기능", inline=False)
            embed.add_field(name="$list  dinoName", value="신청한 공룡 확인기능", inline=False)
            embed.add_field(name="$done  dinoName no", value="공룡 줬으면 삭제하는 기능", inline=False)
            await message.channel.send(embed=embed)

            embed = discord.Embed(title="Command List",  color=0x00ff00)
            
access_token = os.environ["BOT_TOKEN"]
client.run(access_token)
